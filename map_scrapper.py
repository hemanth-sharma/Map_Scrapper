from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
import os

class ManageDriver:
    def __init__(self, url):
        # Load driver
        self.driver = webdriver.Chrome()
        self.driver.maximize_window() # fullscreen
        self.driver.get(url)

    def search_query(self, search_query, searchbox_input_class="searchboxinput"):  
        search_box = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, searchbox_input_class))
        )
        search_box.clear()
        search_box.send_keys(search_query)
        search_box.send_keys(Keys.RETURN)
        time.sleep(10)
    
    def scroll_list(self):
        scrollable_xpath = "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        scrollable_list = WebDriverWait(self.driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, scrollable_xpath))
        )
        previous_height = 0
        
        while True:
            self.driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", scrollable_list)
            time.sleep(2)
            current_height = self.driver.execute_script("return arguments[0].scrollHeight;", scrollable_list)
            
            if current_height == previous_height:
                break
            
            previous_height = current_height

    ### # ## 
    def scroll_and_extract(self):
        extracted_data = []
        processed_elements = set()  # Track already processed elements
        end_marker_class = "PbZDve" # Class indicating the end of the list

        while True:
            # Find all current elements with the target class
            elements = self.driver.find_elements(By.CLASS_NAME, "hfpxzc")

            for element in elements:
                if element not in processed_elements:
                    try:
                        # Scroll to the element
                        self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
                        time.sleep(1)
                        
                        # Extract data
                        aria_label = element.get_attribute("aria-label")
                        href_link = element.get_attribute("href")
                        data = {"aria_label": aria_label, "link": href_link}
                        print(f"Extracted: {data}")
                        extracted_data.append(data)
                        processed_elements.add(element)  # Mark element as processed
                    except Exception as e:
                        print(f"Error processing element: {e}")

            # Check if the end marker is visible
            try:
                end_marker = self.driver.find_element(By.CLASS_NAME, end_marker_class)
                if end_marker.is_displayed():
                    print("End of the list reached.")
                    break
            except:
                pass  # End marker not visible yet

            # Scroll down to load more elements
            self.driver.execute_script("arguments[0].scrollTop += 500;", elements[-1])
            time.sleep(2) 

        return extracted_data
    ## ## ##
    def save_to_excel(self, data, filename):
        os.makedirs("static", exist_ok=True)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Shop Data"

        sheet.cell(row=1, column=1, value="Name")
        sheet.cell(row=1, column=2, value="Link")

        for idx, entry in enumerate(data, start=2):
            sheet.cell(row=idx, column=1, value=entry.get("aria_label"))
            sheet.cell(row=idx, column=2, value=entry.get("link"))

        workbook.save(filename)
        print(f"Data saved to {filename}")

    def extract_shop_data(self):
        # scrollable_list = self.driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]")
        # scrollable_list.find_element("")
        extracted_data = list()
        elements = self.driver.find_elements(By.CLASS_NAME, "hfpxzc")
        
        for index, element in enumerate(elements):
            try:
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
                time.sleep(1)
                aria_label = element.get_attribute("aria-label")   
                href_link = element.get_attribute("href")
                data = {
                    "aria_label": aria_label,
                    "link": href_link
                }
                print(f"Extracted data from element {index + 1}: {data}")
                extracted_data.append(data)

            except Exception as e:
                print("Excepction : ", e)
        
        return extracted_data

    def run(self, search_query):
        try:
            self.search_query(search_query)
            print("Extracting data")
            data = self.scroll_and_extract()
            # self.scroll_list()
            print("Data : ")
            print(data)
            self.save_to_excel(data, "output.xlsx")
            time.sleep(10)
        except Exception as e:
            print("There was some issue during interacting with the driver")
            print("Exception = ", e)

        self.extract_shop_data()

    def terminate(self):
        self.driver.close()


